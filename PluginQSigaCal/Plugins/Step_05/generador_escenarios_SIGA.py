"""
GENERADOR DE ESCENARIOS DINÁMICOS PARA SIGA-CAL
================================================

Script unificado que genera los 4 insumos necesarios para SIGA-CAL:
1. Mapa de prioridad (prioridad.tif)
2. Mapa de transición (transicion.tif)
3. Serie temporal (series_TasRest*.txt)
4. Tabla de parámetros (Tabla_transiciones_idQ.xlsx)

Soporta 4 tipos de escenarios:
- A: Transformación completa (sin predios, sin costos)
- B: Business as Usual (solo degradación)
- C: Con predios priorizados y costos (ACTUAL)
- D: BaU dinámico año a año (FUTURO)

Autor: Sistema automatizado
Fecha: 2025-11-17
"""

import numpy as np
import rasterio
from rasterio.features import geometry_mask
import geopandas as gpd
import pandas as pd
from pathlib import Path
from rasterstats import zonal_stats
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys

# Add parent directory to path to access common_functions
current_dir = Path(__file__).parent
parent_dir = current_dir.parent
if str(parent_dir) not in sys.path:
    sys.path.insert(0, str(parent_dir))


# ==============================================================================
# CONFIGURACIÓN GENERAL
# ==============================================================================

class ConfiguracionEscenario:
    """Configuración centralizada para cada tipo de escenario"""

    def __init__(self, tipo_escenario='C',
                 dict_transformacion=None,
                 valor_fuera_poligono=5,
                 anios_distribucion=15,
                 fecha_inicio='2025-01-01',
                 fecha_fin='2075-12-31',
                 path_bau_anual=None):
        """
        Args:
            tipo_escenario: 'A', 'B', 'C', o 'D'
            dict_transformacion: Diccionario {valor_portafolio: accion_NbS}
                                Ejemplo: {1: 1, 2: 2, 4: 3, 0: 4}
            valor_fuera_poligono: Valor para áreas fuera de polígonos
            anios_distribucion: Años para distribuir implementación (NbS)
            fecha_inicio: Fecha inicio simulación (YYYY-MM-DD)
            fecha_fin: Fecha fin simulación (YYYY-MM-DD)
            path_bau_anual: Path a carpeta con rasters BaU anuales (solo para escenario D)
        """
        self.tipo_escenario = tipo_escenario

        # Rutas base - MUST BE SET BY PLUGIN before use
        self.path_datos = None
        self.path_salida = None

        # Escenario D: BaU dinámico año a año
        self.path_bau_anual = path_bau_anual

        # Parámetros temporales
        self.fecha_inicio = fecha_inicio
        self.fecha_fin = fecha_fin
        self.anios_distribucion = anios_distribucion

        # Diccionario de transformación de portafolio
        if dict_transformacion is None:
            # Default mapping if none provided
            self.dict_transformacion = {
                1: 1,  # Conservación
                2: 2,  # Reforestación activa
                4: 3,  # Regeneración natural
                0: 4   # Sin acción
            }
        else:
            self.dict_transformacion = dict_transformacion

        # Valores especiales
        self.valor_fuera_poligono = valor_fuera_poligono
        self.campo_prioriz = 'prioriz'
        self.valor_prioriz_especial = 999

        # Escenario C: múltiples montos (will be overridden by plugin)
        self.montos_millones = [75, 15, 50, 100, 200]

    def obtener_archivos(self):
        """Retorna diccionario con rutas de archivos según escenario"""
        if self.path_datos is None:
            raise ValueError("path_datos must be set before calling obtener_archivos()")

        base = Path(self.path_datos)

        archivos = {
            'CLC': base / 'CLC.tif',
            'BaU': base / 'CLC_BaU.tif',
            'idQ': base / 'idQ.tif',
        }

        # Portfolio and transitions only for scenarios A and C (NbS scenarios)
        if self.tipo_escenario in ['A', 'C']:
            archivos['Portafolio'] = base / 'Portafolio_NbS_150_clip.tif'
            archivos['Transiciones'] = base / 'transiciones.csv'

        # Scenario C specific files (properties and costs)
        if self.tipo_escenario == 'C':
            archivos['Predios'] = base / 'poligonos_predios.shp'
            archivos['Costos'] = base / 'Portafolio_costo_V2.tif'

        # Scenario B and D don't require Portfolio
        return archivos


# ==============================================================================
# FUNCIONES AUXILIARES
# ==============================================================================

def is_leap_year(year):
    """Verifica si un año es bisiesto"""
    return (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0)


def pixel_area_to_hectares(resolution, area_unit="meters"):
    """Convierte área de píxel a hectáreas"""
    conversion_factor = 0.0001 if area_unit.lower() == "meters" else 1.0
    return resolution * resolution * conversion_factor


def cargar_tabla_transiciones(path_csv):
    """
    Carga tabla de transiciones desde CSV

    CSV debe tener columnas: CLC_inicial, Accion_NbS, CLC_final, ID_transicion

    Retorna diccionario: {(CLC_inicial, Accion_NbS): (CLC_final, ID_transicion, Accion_NbS)}
    """
    df = pd.read_csv(path_csv)

    tabla = {}
    for _, row in df.iterrows():
        clave = (int(row['CLC_inicial']), int(row['Accion_NbS']))
        valor = (int(row['CLC_final']), int(row['ID_transicion']), int(row['Accion_NbS']))
        tabla[clave] = valor

    print(f"✓ Tabla de transiciones cargada: {len(df)} transiciones")
    return tabla


# ==============================================================================
# GENERADOR DE ARCHIVO 02_Restauracion.xlsx (ÚNICO PARA TODOS LOS ESCENARIOS)
# ==============================================================================

def generar_02_restauracion_unico(config):
    """
    Genera archivo único 02_Restauracion.xlsx con parámetros de calibración

    Este archivo es común para todos los escenarios ya que contiene
    los parámetros de las transiciones, no depende del presupuesto.
    """
    print("\n" + "=" * 80)
    print("GENERANDO ARCHIVO 02_Restauracion.xlsx (ÚNICO)")
    print("=" * 80)

    # Encontrar una tabla de transiciones generada (usar cualquier escenario)
    outputs_path = Path(config.path_salida)
    tabla_encontrada = None

    for subdir in outputs_path.iterdir():
        if subdir.is_dir():
            tabla_path = subdir / "Tabla_transiciones_idQ.xlsx"
            if tabla_path.exists():
                tabla_encontrada = tabla_path
                print(f"Usando tabla de: {subdir.name}/")
                break

    if not tabla_encontrada:
        print("⚠ No se encontró tabla generada. Saltando generación de 02_Restauracion.xlsx")
        return

    # Leer tablas
    df_trans_generada = pd.read_excel(tabla_encontrada)

    # Load original transitions CSV only if it exists (scenarios A, C)
    df_trans_original = None
    transiciones_csv_path = Path(config.path_datos) / 'transiciones.csv'
    if transiciones_csv_path.exists():
        df_trans_original = pd.read_csv(transiciones_csv_path)
        print(f"Transiciones originales cargadas: {len(df_trans_original)}")
    else:
        print("No transiciones.csv found (expected for scenarios B and D)")

    print(f"Transiciones en tabla generada: {len(df_trans_generada)}")

    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # === HOJA 1: Transiciones ===
    ws_trans = wb.create_sheet("Transiciones")

    headers = ["ID_Transicion", "CLC_Inicial", "CLC_Final", "Accion_NbS",
               "idQ", "deltatm", "Nombre_Transicion", "Tipo"]

    for col, header in enumerate(headers, start=1):
        cell = ws_trans.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for _, trans in df_trans_generada.iterrows():
        trans_id = int(trans['transicion'])
        clc_end = int(trans['CLC End'])
        accion = int(trans['accion'])
        idq = int(trans['idQ'])
        deltatm = int(trans['deltatm'])

        # Buscar info en CSV original
        clc_inicial = -9999
        nombre = "Transición automática"
        tipo = "Degradación" if accion == 7 else "NbS"

        if accion != 7 and df_trans_original is not None:
            match = df_trans_original[df_trans_original['ID_transicion'] == trans_id]
            if not match.empty:
                clc_inicial = int(match.iloc[0]['CLC_inicial'])
                nombre = match.iloc[0].get('Nombre_Transicion', f'Transición {trans_id}')
                tipo = "NbS"

        ws_trans.cell(row=row, column=1, value=trans_id)
        ws_trans.cell(row=row, column=2, value=clc_inicial)
        ws_trans.cell(row=row, column=3, value=clc_end)
        ws_trans.cell(row=row, column=4, value=accion)
        ws_trans.cell(row=row, column=5, value=idq)
        ws_trans.cell(row=row, column=6, value=deltatm)
        ws_trans.cell(row=row, column=7, value=nombre)
        ws_trans.cell(row=row, column=8, value=tipo)
        row += 1

    # Ajustar anchos
    ws_trans.column_dimensions['A'].width = 15
    ws_trans.column_dimensions['G'].width = 50

    print(f"✓ Hoja 'Transiciones': {row-1} transiciones")

    # === HOJA 2: Acciones NbS ===
    ws_acciones = wb.create_sheet("Acciones_NbS")

    headers_acc = ["ID_Accion", "Nombre", "Descripcion", "Tipo"]
    for col, header in enumerate(headers_acc, start=1):
        cell = ws_acciones.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    acciones = [
        (1, "Conservación", "Protección de ecosistemas", "Positiva"),
        (2, "Reforestación", "Plantación activa", "Positiva"),
        (3, "Regeneración", "Recuperación asistida", "Positiva"),
        (4, "Sin Acción", "Mantenimiento", "Neutra"),
        (5, "Fuera Intervención", "Fuera de predios", "Neutra"),
        (7, "Degradación", "Cambio negativo", "Negativa")
    ]

    for row, (id_acc, nombre, desc, tipo) in enumerate(acciones, start=2):
        ws_acciones.cell(row=row, column=1, value=id_acc)
        ws_acciones.cell(row=row, column=2, value=nombre)
        ws_acciones.cell(row=row, column=3, value=desc)
        ws_acciones.cell(row=row, column=4, value=tipo)

    ws_acciones.column_dimensions['B'].width = 20
    ws_acciones.column_dimensions['C'].width = 40

    print(f"✓ Hoja 'Acciones_NbS': {len(acciones)} acciones")

    # === HOJA 3: Parámetros Temporales ===
    ws_temp = wb.create_sheet("Parametros")

    ws_temp.cell(row=1, column=1, value="Parámetro")
    ws_temp.cell(row=1, column=2, value="Valor")

    for col in [1, 2]:
        cell = ws_temp.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    params = [
        ("Fecha_Inicio", config.fecha_inicio),
        ("Fecha_Fin", config.fecha_fin),
        ("Años_Implementacion", str(config.anios_distribucion)),
        ("Sistema_CLC", "Corine Land Cover adaptado"),
        ("Modelo", "SIGA-CAL")
    ]

    for row, (param, valor) in enumerate(params, start=2):
        ws_temp.cell(row=row, column=1, value=param)
        ws_temp.cell(row=row, column=2, value=valor)

    ws_temp.column_dimensions['A'].width = 25
    ws_temp.column_dimensions['B'].width = 30

    print(f"✓ Hoja 'Parametros': {len(params)} parámetros")

    # Guardar en Outputs (no sobreescribir ejemplo)
    path_salida = Path(config.path_salida) / "02_Restauracion_generado.xlsx"
    wb.save(path_salida)

    print(f"\n✓ ARCHIVO GENERADO: {path_salida}")
    print("=" * 80)

    # También generar archivo de áreas acumuladas (complemento)
    generar_areas_acumuladas(config, df_trans_generada)


def generar_areas_acumuladas(config, df_trans):
    """
    Genera archivo complementario con áreas acumuladas por presupuesto
    """
    print("\nGenerando archivo de áreas acumuladas...")

    # Buscar archivos de áreas generados
    outputs_path = Path(config.path_salida)
    datos_areas = []

    for subdir in sorted(outputs_path.iterdir()):
        if subdir.is_dir() and subdir.name.isdigit():
            areas_file = subdir / 'Areas_accion.xlsx'
            if areas_file.exists():
                df_area = pd.read_excel(areas_file)
                df_area['Presupuesto_M'] = int(subdir.name)
                datos_areas.append(df_area)

    if not datos_areas:
        print("⚠ No se encontraron archivos de áreas")
        return

    # Consolidar
    df_consolidado = pd.concat(datos_areas, ignore_index=True)

    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Areas_por_Presupuesto"

    # Headers
    headers = ["Presupuesto_M", "Clase", "Area_Hectareas", "Area_m2"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # Datos
    row = 2
    for _, data in df_consolidado.iterrows():
        ws.cell(row=row, column=1, value=data['Presupuesto_M'])
        ws.cell(row=row, column=2, value=int(data['Clase']))
        ws.cell(row=row, column=3, value=data['Area_Hectareas'])
        ws.cell(row=row, column=4, value=data['Area_m2'])
        row += 1

    # Guardar
    path_areas = Path(config.path_salida) / "Consolidado_Areas_Presupuestos.xlsx"
    wb.save(path_areas)

    print(f"✓ Áreas consolidadas: {path_areas}")


# ==============================================================================
# ESCENARIO C: CON PREDIOS Y COSTOS (FUNCIÓN PRINCIPAL ACTUAL)
# ==============================================================================

def generar_escenario_C(config):
    """
    ESCENARIO C: Con predios priorizados y filtro por costos

    Genera múltiples escenarios según lista de presupuestos
    """
    print("=" * 80)
    print("EJECUTANDO ESCENARIO C: CON PREDIOS PRIORIZADOS Y COSTOS")
    print("=" * 80)

    archivos = config.obtener_archivos()

    # Validar archivos
    for nombre, path in archivos.items():
        if not Path(path).exists():
            raise FileNotFoundError(f"Archivo no encontrado: {nombre} -> {path}")

    # Cargar tabla de transiciones
    tabla_transiciones = cargar_tabla_transiciones(str(archivos['Transiciones']))

    # Procesar cada monto
    for monto in config.montos_millones:
        print("\n" + "=" * 80)
        print(f"PROCESANDO PRESUPUESTO: ${monto}M USD")
        print("=" * 80)

        # 1. Seleccionar predios por presupuesto
        path_shp_filtrado = seleccionar_predios_por_presupuesto(
            path_shp=str(archivos['Predios']),
            path_raster_costos=str(archivos['Costos']),
            monto_millones=monto,
            path_out=config.path_salida,
            campo_prioriz=config.campo_prioriz,
            valor_especial=config.valor_prioriz_especial
        )

        # Directorio de salida para este monto
        dir_salida = Path(config.path_salida) / str(monto)
        dir_salida.mkdir(parents=True, exist_ok=True)

        # 2. Generar mapa de prioridad
        generar_mapa_prioridad(
            path_raster_base=str(archivos['Portafolio']),
            path_shp=path_shp_filtrado,
            path_salida=str(dir_salida / 'prioridad.tif')
        )

        # 3. Generar portafolio de predios
        path_portafolio_predios = generar_portafolio_predios(
            path_raster=str(archivos['Portafolio']),
            dict_transformacion=config.dict_transformacion,
            path_salida=str(dir_salida),
            path_shp=path_shp_filtrado,
            valor_fuera_poligono=config.valor_fuera_poligono
        )

        # 4. Generar mapa de transición
        generar_mapa_transicion(
            path_clc=str(archivos['CLC']),
            path_bau=str(archivos['BaU']),
            path_idq=str(archivos['idQ']),
            path_portafolio=path_portafolio_predios,
            tabla_transiciones=tabla_transiciones,
            path_salida=str(dir_salida)
        )

        # 5. Generar serie temporal
        generar_serie_temporal(
            path_raster=path_portafolio_predios,
            start_date=config.fecha_inicio,
            end_date=config.fecha_fin,
            distribucion_anios=config.anios_distribucion,
            path_salida=str(dir_salida),
            nombre_escenario=str(monto)
        )

        print(f"\n✓ Escenario ${monto}M USD COMPLETADO")
        print(f"  Resultados en: {dir_salida}")

    print("\n" + "=" * 80)
    print(f"✓ ESCENARIO C COMPLETADO: {len(config.montos_millones)} presupuestos procesados")
    print("=" * 80)

    # Generar archivo único 02_Restauracion.xlsx
    generar_02_restauracion_unico(config)


def generar_prioridad_raster(path_raster_base, path_shp, path_salida):
    """
    Genera raster de prioridad.
    Si path_shp es None, prioriza todo el raster sin usar shapefile.
    """
    with rasterio.open(path_raster_base) as src:
        raster_base = src.read(1)
        meta = src.meta.copy()
        shape = raster_base.shape
        transform = src.transform
        nodata = src.nodata

    raster_prioridad = np.zeros(shape, dtype='float32')
    pixel_id = 1

    if path_shp is None:
        # Sin shapefile: priorizar todo el raster
        mascara_valida = np.ones_like(raster_base, dtype=bool) if nodata is None else raster_base != nodata
        rows, cols = np.where(mascara_valida)
        for i in range(len(rows)):
            raster_prioridad[rows[i], cols[i]] = pixel_id
            pixel_id += 1
    else:
        # Con shapefile (para otros escenarios)
        gdf = gpd.read_file(path_shp)
        gdf = gdf[gdf['prioriz'] != 0].sort_values(by='prioriz', ascending=True)

        for _, row in gdf.iterrows():
            geom = [row.geometry]
            mask = geometry_mask(geom, transform=transform, invert=True, out_shape=shape)
            rows_mask, cols_mask = np.where(mask)
            for i in range(len(rows_mask)):
                raster_prioridad[rows_mask[i], cols_mask[i]] = pixel_id
                pixel_id += 1

    # Guardar
    path_out = Path(path_salida) / 'prioridad.tif'
    with rasterio.open(path_out, 'w', **meta) as dst:
        dst.write(raster_prioridad, 1)


def generar_portafolio_sin_predios(path_portafolio, dict_transformacion, valor_fuera_poligono, path_salida):
    """
    Transforma el portafolio original aplicando dict_transformacion.
    Sin usar shapefile de predios.
    """
    with rasterio.open(path_portafolio) as src:
        raster = src.read(1)
        meta = src.meta.copy()

    resultado = np.full_like(raster, valor_fuera_poligono, dtype=np.int32)

    for valor_original, valor_nuevo in dict_transformacion.items():
        mascara = raster == valor_original
        resultado[mascara] = valor_nuevo

    # Guardar
    path_out = Path(path_salida) / 'Portafolio_Predios.tif'
    with rasterio.open(path_out, 'w', **meta) as dst:
        dst.write(resultado, 1)


# ==============================================================================
# ESCENARIO A: TRANSFORMACIÓN COMPLETA (SIN PREDIOS)
# ==============================================================================

def generar_escenario_A(config):
    """
    ESCENARIO A: Transformación completa del portafolio

    Sin shapefile de predios, sin restricción de costos.
    Implementa TODO el portafolio NbS con transiciones.
    """
    print("=" * 80)
    print("EJECUTANDO ESCENARIO A: TRANSFORMACIÓN COMPLETA")
    print("=" * 80)

    archivos = config.obtener_archivos()
    dir_salida = Path(config.path_salida) / 'Escenario_A'
    dir_salida.mkdir(parents=True, exist_ok=True)

    # Cargar tabla de transiciones
    tabla_transiciones = cargar_tabla_transiciones(str(archivos['Transiciones']))
    print(f"✓ Tabla de transiciones cargada: {len(tabla_transiciones)} transiciones")

    # Portafolio original
    path_portafolio_original = archivos['Portafolio']

    # 1. Generar prioridad desde portafolio (sin shapefile)
    print("\n1. Generando mapa de prioridad...")
    generar_prioridad_raster(
        path_raster_base=str(path_portafolio_original),
        path_shp=None,
        path_salida=str(dir_salida)
    )

    # 2. Generar Portafolio_Predios.tif (aplicar transformación NbS)
    print("\n2. Generando Portafolio_Predios.tif...")
    generar_portafolio_sin_predios(
        path_portafolio=str(path_portafolio_original),
        dict_transformacion=config.dict_transformacion,
        valor_fuera_poligono=config.valor_fuera_poligono,
        path_salida=str(dir_salida)
    )

    path_portafolio_predios = dir_salida / 'Portafolio_Predios.tif'

    # 3. Generar mapa de transición (con NbS y degradaciones)
    print("\n3. Generando mapa de transición...")
    generar_mapa_transicion(
        path_clc=str(archivos['CLC']),
        path_bau=str(archivos['BaU']),
        path_idq=str(archivos['idQ']),
        path_portafolio=str(path_portafolio_predios),
        tabla_transiciones=tabla_transiciones,
        path_salida=str(dir_salida)
    )

    # 4. Generar serie temporal
    print("\n4. Generando serie temporal...")
    generar_serie_temporal(
        path_raster=str(path_portafolio_predios),
        start_date=config.fecha_inicio,
        end_date=config.fecha_fin,
        distribucion_anios=config.anios_distribucion,
        path_salida=str(dir_salida),
        nombre_escenario='A'
    )

    print("\n✓ ESCENARIO A COMPLETADO")
    print(f"  Resultados en: {dir_salida}")

    # Generar archivo único 02_Restauracion.xlsx
    generar_02_restauracion_unico(config)


# ==============================================================================
# ESCENARIO B: BUSINESS AS USUAL (SOLO DEGRADACIÓN)
# ==============================================================================

def generar_escenario_B(config):
    """
    ESCENARIO B: Business as Usual (No NbS interventions)

    This scenario doesn't require Portfolio NbS raster. Uses CLC as base.
    1. Generar prioridad desde CLC (all valid areas have same priority)
    2. Generar Portafolio_Predios.tif desde CLC (all category 5 = no action)
    3. Generar mapa de transición CLC → BaU (degradations only)
    4. Generar serie temporal (only column 5 - degradation/no action)
    """
    print("=" * 80)
    print("EJECUTANDO ESCENARIO B: BUSINESS AS USUAL")
    print("=" * 80)

    archivos = config.obtener_archivos()
    dir_salida = Path(config.path_salida) / 'Escenario_B_BaU'
    dir_salida.mkdir(parents=True, exist_ok=True)

    # Use CLC as base (Scenario B doesn't require Portfolio)
    path_clc_base = archivos['CLC']

    # 1. Generar prioridad desde CLC (all valid areas have priority = 1)
    print("\n1. Generando mapa de prioridad desde CLC...")
    generar_prioridad_raster(
        path_raster_base=str(path_clc_base),
        path_shp=None,
        path_salida=str(dir_salida)
    )

    # 2. Generar Portafolio_Predios.tif from CLC (para BaU: todo es categoría 5)
    print("\n2. Generando Portafolio_Predios.tif desde CLC (todo categoría 5 para BaU)...")
    with rasterio.open(path_clc_base) as src:
        clc_raster = src.read(1)
        meta_port = src.meta.copy()

    # Para BaU: todo lo válido de CLC se convierte a categoría 5 (no action = degradation only)
    portafolio_bau = np.where(clc_raster != -9999, 5, -9999).astype(np.int32)

    path_portafolio_predios = dir_salida / 'Portafolio_Predios.tif'
    with rasterio.open(path_portafolio_predios, 'w', **meta_port) as dst:
        dst.write(portafolio_bau, 1)

    print(f"  ✓ Portafolio_Predios.tif generado (todo categoría 5)")

    # 3. Generar mapa de transición (CLC → BaU con máscara de Portafolio_Predios)
    print("\n3. Generando mapa de transición...")

    # Leer rasters (usando Portafolio_Predios como máscara)
    with rasterio.open(archivos['CLC']) as src:
        CLC = src.read(1)
        meta = src.meta.copy()
        resolution = src.res[0]

    with rasterio.open(archivos['BaU']) as src:
        BaU = src.read(1)

    with rasterio.open(archivos['idQ']) as src:
        IDQ = src.read(1)

    with rasterio.open(path_portafolio_predios) as src:
        Portafolio = src.read(1)

    print(f"  Dimensiones: {CLC.shape}")
    print(f"  Resolución: {resolution} m")

    # Generar mapa de transición (solo degradaciones dentro del portafolio)
    resultado = np.full_like(CLC, 1, dtype=np.int32)
    transicion_info = []
    degradaciones = {}
    next_trans_id = 2

    sin_cambio = 0
    con_degradacion = 0

    rows, cols = CLC.shape

    print("  Procesando píxeles...")
    for i in range(rows):
        for j in range(cols):
            clc = CLC[i, j]
            bau = BaU[i, j]
            idq = IDQ[i, j]
            portafolio = Portafolio[i, j]

            # IMPORTANTE: Usar portafolio como máscara (igual que código original)
            if clc == -9999 or bau == -9999 or idq == -9999 or portafolio == -9999:
                continue

            if clc == bau:
                resultado[i, j] = 1
                sin_cambio += 1
            else:
                clave_deg = (int(clc), int(bau))
                if clave_deg not in degradaciones:
                    degradaciones[clave_deg] = next_trans_id
                    next_trans_id += 1

                trans_id = degradaciones[clave_deg]
                resultado[i, j] = trans_id
                transicion_info.append((trans_id, bau, 1, idq, 7))
                con_degradacion += 1

        if (i + 1) % 100 == 0:
            print(f"    Fila {i+1}/{rows} ({(i+1)/rows*100:.1f}%)")

    print(f"  Píxeles sin cambio: {sin_cambio}")
    print(f"  Píxeles con degradación (con máscara portafolio): {con_degradacion}")

    # Calcular área de degradación
    area_pixel = resolution * resolution
    area_con_mascara = con_degradacion * area_pixel

    # PRUEBA: Calcular sin máscara de portafolio (solo validar nodata de CLC, BaU, idQ)
    print(f"\n  Calculando área SIN máscara de portafolio...")
    degr_sin_mascara = 0
    for i in range(rows):
        for j in range(cols):
            if CLC[i, j] != -9999 and BaU[i, j] != -9999 and IDQ[i, j] != -9999:
                if CLC[i, j] != BaU[i, j]:
                    degr_sin_mascara += 1

    area_sin_mascara = degr_sin_mascara * area_pixel

    print(f"\n  === COMPARACIÓN DE ÁREAS ===")
    print(f"  CON máscara portafolio:")
    print(f"    Píxeles: {con_degradacion:,}")
    print(f"    Área: {area_con_mascara:.2e} m²")
    print(f"    Área diaria (1 año): {area_con_mascara/365:.2e} m²/día")

    print(f"\n  SIN máscara portafolio:")
    print(f"    Píxeles: {degr_sin_mascara:,}")
    print(f"    Área: {area_sin_mascara:.2e} m²")
    print(f"    Área diaria (1 año): {area_sin_mascara/365:.2e} m²/día")

    # Usar el área SIN máscara (todo el raster válido)
    area_degradacion_total = area_sin_mascara
    print(f"\n  >>> USANDO: {area_degradacion_total:.2e} m² <<<")

    # Guardar raster
    with rasterio.open(str(dir_salida / 'transicion.tif'), 'w', **meta) as dst:
        dst.write(resultado, 1)

    # Generar tabla de transiciones
    df_trans = pd.DataFrame(transicion_info, columns=["transicion", "CLC End", "deltatm", "idQ", "accion"])
    df_trans = df_trans.drop_duplicates()

    fila_trans_1 = pd.DataFrame([{"transicion": 1, "CLC End": -9999, "deltatm": 0, "idQ": 1, "accion": 1}])
    df_final = pd.concat([fila_trans_1, df_trans], ignore_index=True)
    df_final = df_final.sort_values(by="transicion").reset_index(drop=True)
    df_final.to_excel(str(dir_salida / 'Tabla_transiciones_idQ.xlsx'), index=False)

    # 4. Generar serie temporal desde Portafolio_Predios (solo columna 5)
    print("\n4. Generando serie temporal desde Portafolio_Predios.tif...")
    generar_serie_temporal(
        path_raster=str(path_portafolio_predios),
        start_date=config.fecha_inicio,
        end_date=config.fecha_fin,
        distribucion_anios=config.anios_distribucion,
        path_salida=str(dir_salida),
        nombre_escenario='BaU'
    )

    print("\n✓ ESCENARIO B COMPLETADO")
    print(f"  Resultados en: {dir_salida}")

    # Generar archivo único 02_Restauracion.xlsx (incluye degradaciones)
    generar_02_restauracion_unico(config)


# ==============================================================================
# ESCENARIO D: BAU DINÁMICO AÑO A AÑO
# ==============================================================================

def generar_escenario_D(config):
    """
    ESCENARIO D: BaU dinámico año a año

    En lugar de un solo BaU fijo, usa rasters BaU específicos por año.
    Calcula degradación año a año: año1 (CLC→BaU_2025), año2 (BaU_2025→BaU_2026), etc.

    Requiere: config.path_bau_anual con rasters BaU_YYYY.tif
    """
    print("=" * 80)
    print("EJECUTANDO ESCENARIO D: BAU DINÁMICO AÑO A AÑO")
    print("=" * 80)

    if not config.path_bau_anual:
        raise ValueError("Escenario D requiere path_bau_anual configurado")

    archivos = config.obtener_archivos()
    dir_salida = Path(config.path_salida) / 'Escenario_D_BaU_Dinamico'
    dir_salida.mkdir(parents=True, exist_ok=True)

    # Portafolio original
    path_portafolio_original = Path(config.path_datos) / 'Portafolio_NbS_150_clip.tif'

    # 1. Generar prioridad
    print("\n1. Generando mapa de prioridad...")
    generar_prioridad_raster(
        path_raster_base=str(path_portafolio_original),
        path_shp=None,
        path_salida=str(dir_salida)
    )

    # 2. Generar Portafolio_Predios.tif (todo categoría 5 para BaU)
    print("\n2. Generando Portafolio_Predios.tif...")
    with rasterio.open(path_portafolio_original) as src:
        portafolio_original = src.read(1)
        meta_port = src.meta.copy()

    portafolio_bau = np.where(portafolio_original != -9999, 5, -9999).astype(np.int32)
    path_portafolio_predios = dir_salida / 'Portafolio_Predios.tif'
    with rasterio.open(path_portafolio_predios, 'w', **meta_port) as dst:
        dst.write(portafolio_bau, 1)
    print(f"  ✓ Portafolio_Predios.tif generado")

    # 3. Generar transiciones y series año a año
    print("\n3. Generando transiciones año a año...")

    # Obtener años del periodo
    start_year = pd.to_datetime(config.fecha_inicio).year
    end_year = pd.to_datetime(config.fecha_fin).year

    # Leer rasters base
    with rasterio.open(archivos['CLC']) as src:
        CLC_inicial = src.read(1)
        meta = src.meta.copy()
        resolution = src.res[0]

    with rasterio.open(archivos['idQ']) as src:
        IDQ = src.read(1)

    with rasterio.open(path_portafolio_predios) as src:
        Portafolio = src.read(1)

    # Generar transiciones y series año a año
    areas_anuales = generar_transiciones_bau_anual(
        CLC_inicial=CLC_inicial,
        IDQ=IDQ,
        Portafolio=Portafolio,
        path_bau_anual=config.path_bau_anual,
        start_year=start_year,
        end_year=end_year,
        resolution=resolution,
        meta=meta,
        dir_salida=dir_salida
    )

    # 4. Generar serie temporal consolidada
    print("\n4. Generando serie temporal consolidada...")
    generar_serie_temporal_bau_dinamico(
        areas_anuales=areas_anuales,
        config=config,
        path_salida=str(dir_salida)
    )

    print("\n✓ ESCENARIO D COMPLETADO")
    print(f"  Resultados en: {dir_salida}")

    # Generar archivo único 02_Restauracion.xlsx
    generar_02_restauracion_unico(config)


def generar_transiciones_bau_anual(CLC_inicial, IDQ, Portafolio, path_bau_anual,
                                     start_year, end_year, resolution, meta, dir_salida):
    """
    Genera transiciones año a año para BaU dinámico

    Returns:
        dict: {año: area_degradacion_m2}
    """
    path_bau_dir = Path(path_bau_anual)
    areas_anuales = {}

    # Estado actual (comienza con CLC inicial)
    estado_actual = CLC_inicial.copy()

    area_pixel = resolution * resolution

    # Procesar cada año
    for año in range(start_year, end_year + 1):
        # Buscar archivo BaU para este año
        path_bau_año = path_bau_dir / f'BaU_{año}.tif'

        if not path_bau_año.exists():
            print(f"  ⚠ No existe BaU_{año}.tif, saltando...")
            areas_anuales[año] = 0.0
            continue

        print(f"  Procesando año {año}...")

        # Leer BaU del año
        with rasterio.open(path_bau_año) as src:
            BaU_año = src.read(1)

        # Calcular degradación para este año
        degradacion_año = 0
        rows, cols = estado_actual.shape

        for i in range(rows):
            for j in range(cols):
                estado = estado_actual[i, j]
                bau = BaU_año[i, j]
                port = Portafolio[i, j]
                idq = IDQ[i, j]

                # Validar
                if estado == -9999 or bau == -9999 or port == -9999 or idq == -9999:
                    continue

                # Si hay cambio = degradación
                if estado != bau:
                    degradacion_año += 1
                    estado_actual[i, j] = bau  # Actualizar estado

        # Calcular área
        area_degradacion = degradacion_año * area_pixel
        areas_anuales[año] = area_degradacion

        print(f"    Píxeles degradados: {degradacion_año:,}")
        print(f"    Área: {area_degradacion:,.2f} m² ({area_degradacion/10000:,.2f} ha)")

    # Guardar estado final como transicion.tif
    with rasterio.open(dir_salida / 'transicion_final.tif', 'w', **meta) as dst:
        dst.write(estado_actual, 1)

    return areas_anuales


def generar_serie_temporal_bau_dinamico(areas_anuales, config, path_salida):
    """
    Genera serie temporal con degradación año a año

    Args:
        areas_anuales: dict {año: area_m2}
    """
    # Generar fechas
    fechas = pd.date_range(start=config.fecha_inicio, end=config.fecha_fin)
    df = pd.DataFrame(index=fechas)

    df['Año'] = df.index.year
    df['Mes'] = df.index.month
    df['Día'] = df.index.day

    # Inicializar categorías
    categorias = [1, 2, 3, 4, 5]
    for cat in categorias:
        df[cat] = 0.0

    # Distribuir área de cada año en los 365 días de ese año
    for año, area_total in areas_anuales.items():
        # Filtrar días del año
        mask_año = df['Año'] == año
        dias_año = mask_año.sum()

        if dias_año > 0:
            area_diaria = area_total / dias_año
            df.loc[mask_año, 5] = area_diaria

    # Reordenar
    df = df[['Año', 'Mes', 'Día', 1, 2, 3, 4, 5]]

    # Exportar CSV
    df.to_csv(Path(path_salida) / 'Serie_diaria_areas.csv', index=False)

    # Exportar SIGA
    try:
        from common_functions.ficheros import Ficheros

        df_meta = pd.DataFrame({
            "codigo": categorias,
            "x": [-9999] * len(categorias),
            "y": [-9999] * len(categorias),
            "z": [-9999] * len(categorias)
        })

        dir_series = Path(path_salida) / 'series'
        dir_series.mkdir(exist_ok=True)

        df_siga = df[categorias].copy()

        ficheros = Ficheros(ruta_cuenca_SIGA=str(path_salida) + '/')
        ficheros.ExportarSeriesSIGA(
            df=df_siga,
            df_metadatos=df_meta,
            nombre_archivo_serie='/series_TasRestBaU_Dinamico.txt'
        )
        print(f"  ✓ Serie SIGA exportada")
    except ImportError:
        print(f"  ⚠ Módulo 'ficheros' no disponible - solo CSV exportado")
    except Exception as e:
        print(f"  ⚠ Error al exportar SIGA: {e}")

    print(f"  ✓ Serie temporal dinámica generada")


# ==============================================================================
# FUNCIONES CORE
# ==============================================================================

def seleccionar_predios_por_presupuesto(path_shp, path_raster_costos, monto_millones,
                                        path_out, campo_prioriz='prioriz', valor_especial=999):
    """
    Selecciona predios según presupuesto acumulado

    Args:
        path_shp: Shapefile con predios y campo de priorización
        path_raster_costos: Raster con costos por píxel (USD)
        monto_millones: Presupuesto en millones de USD
        path_out: Directorio de salida
        campo_prioriz: Nombre del campo de priorización
        valor_especial: Valor para predios que siempre se incluyen (ej. 999)

    Returns:
        Ruta al shapefile generado
    """
    print(f"\n1. Seleccionando predios para ${monto_millones}M USD...")

    df_shp = gpd.read_file(path_shp)

    # Separar predios especiales y priorizados
    df_999 = df_shp[df_shp[campo_prioriz] == valor_especial].copy()
    df_shp = df_shp[(df_shp[campo_prioriz] != 0) & (df_shp[campo_prioriz] != valor_especial)].copy()
    df_shp = df_shp.sort_values(by=campo_prioriz, ascending=True)

    print(f"  Predios a priorizar: {len(df_shp)}")
    print(f"  Predios especiales (999): {len(df_999)}")

    # Calcular costos
    stats_costos = zonal_stats(df_shp, path_raster_costos, stats='sum')
    df_costos = pd.DataFrame(stats_costos, index=df_shp.index)
    df_costos['Acumulado'] = df_costos['sum'].cumsum()

    df_shp = df_shp.merge(df_costos, left_index=True, right_index=True)

    # Filtrar por monto
    df_filtrado = df_shp[df_shp['Acumulado'] < 1e6 * monto_millones]

    print(f"  Predios seleccionados: {len(df_filtrado)}")

    # Unir con predios especiales
    df_resultado = pd.concat([df_filtrado, df_999], ignore_index=True)

    print(f"  Total predios: {len(df_resultado)}")

    # Guardar
    nombre_escenario = f"Scenario_{monto_millones}M"
    dir_out = Path(path_out) / str(monto_millones)
    dir_out.mkdir(parents=True, exist_ok=True)

    path_out_shp = dir_out / f"{nombre_escenario}.shp"
    df_resultado.to_file(path_out_shp, driver='ESRI Shapefile')

    print(f"  ✓ Shapefile guardado: {path_out_shp.name}")

    return str(path_out_shp)


def generar_mapa_prioridad(path_raster_base, path_shp, path_salida):
    """
    Genera mapa de prioridad

    Args:
        path_raster_base: Raster base (portafolio)
        path_shp: Shapefile con predios (None = prioridad uniforme)
        path_salida: Ruta de salida del raster
    """
    print(f"\n2. Generando mapa de prioridad...")

    with rasterio.open(path_raster_base) as src:
        raster_base = src.read(1)
        meta = src.meta.copy()
        shape = raster_base.shape
        transform = src.transform
        nodata = src.nodata

    raster_prioridad = np.zeros(shape, dtype='float32')
    pixel_id = 1

    if path_shp and Path(path_shp).exists():
        print("  Priorizando con shapefile...")
        gdf = gpd.read_file(path_shp)

        if 'prioriz' in gdf.columns:
            gdf = gdf[gdf['prioriz'] != 0].sort_values(by='prioriz', ascending=True)

        # Asignar prioridades por polígono
        for _, row in gdf.iterrows():
            geom = [row.geometry]
            mask = geometry_mask(geom, transform=transform, invert=True, out_shape=shape)
            rows, cols = np.where(mask)
            for i in range(len(rows)):
                raster_prioridad[rows[i], cols[i]] = pixel_id
                pixel_id += 1

        # Resto fuera de polígonos
        mask_outside = geometry_mask(gdf.geometry, transform=transform, invert=False, out_shape=shape)
        rows, cols = np.where(mask_outside & (raster_base != nodata))
        for i in range(len(rows)):
            if raster_prioridad[rows[i], cols[i]] == 0:
                raster_prioridad[rows[i], cols[i]] = pixel_id
                pixel_id += 1
    else:
        print("  Priorización uniforme (sin shapefile)...")
        mascara_valida = np.ones_like(raster_base, dtype=bool) if nodata is None else raster_base != nodata
        rows, cols = np.where(mascara_valida)
        for i in range(len(rows)):
            raster_prioridad[rows[i], cols[i]] = pixel_id
            pixel_id += 1

    print(f"  Píxeles priorizados: {pixel_id - 1}")

    # Guardar
    meta.update({'dtype': 'float32', 'nodata': 0})
    Path(path_salida).parent.mkdir(parents=True, exist_ok=True)

    with rasterio.open(path_salida, 'w', **meta) as dst:
        dst.write(raster_prioridad, 1)

    print(f"  ✓ Prioridad guardada: {Path(path_salida).name}")


def generar_portafolio_predios(path_raster, dict_transformacion, path_salida,
                                path_shp=None, valor_fuera_poligono=7):
    """
    Genera raster de portafolio con predios (Transición Parte A)

    Args:
        path_raster: Raster de portafolio original
        dict_transformacion: {valor_original: nuevo_valor}
        path_salida: Directorio de salida
        path_shp: Shapefile de predios (opcional)
        valor_fuera_poligono: Valor para áreas fuera de polígonos

    Returns:
        Ruta al raster generado
    """
    print(f"\n3. Generando portafolio de predios...")

    with rasterio.open(path_raster) as src:
        raster = src.read(1)
        meta = src.meta.copy()
        transform = src.transform
        resolution = src.res[0]
        shape = raster.shape

    # Aplicar transformación
    raster_por = raster.copy()
    for clase_i, clase_f in dict_transformacion.items():
        raster_por[raster == clase_i] = clase_f

    # Aplicar máscara de shapefile
    if path_shp and Path(path_shp).exists():
        df_shp = gpd.read_file(path_shp)
        mask = geometry_mask(df_shp.geometry, transform=transform, invert=False, out_shape=shape)
        raster_por[mask] = valor_fuera_poligono
        print(f"  Shapefile aplicado: {len(df_shp)} predios")
    else:
        print("  Sin shapefile → TODO degradación")
        raster_por[:] = valor_fuera_poligono

    raster_por = np.round(raster_por).astype('float32')

    # Guardar
    path_out = Path(path_salida) / 'Portafolio_Predios.tif'
    with rasterio.open(path_out, 'w', **meta) as dst:
        dst.write(raster_por, 1)

    # Tabla de áreas
    resumen = {"Clase": [], "Area_Hectareas": [], "Area_m2": []}
    for clase in np.unique(raster_por):
        if clase == 0:
            continue
        n_pixels = np.sum(raster_por == clase)
        area_ha = pixel_area_to_hectares(resolution) * n_pixels
        resumen["Clase"].append(int(clase))
        resumen["Area_Hectareas"].append(area_ha)
        resumen["Area_m2"].append(area_ha * 10000)

    df = pd.DataFrame(resumen)
    df.to_excel(Path(path_salida) / 'Areas_accion.xlsx', index=False)

    print(f"  ✓ Portafolio guardado: {path_out.name}")

    return str(path_out)


def generar_mapa_transicion(path_clc, path_bau, path_idq, path_portafolio,
                            tabla_transiciones, path_salida):
    """
    Genera mapa de transición final (Transición Parte B)

    Args:
        path_clc: Raster de cobertura actual
        path_bau: Raster de escenario BaU
        path_idq: Raster de índice de calidad
        path_portafolio: Raster de portafolio con predios
        tabla_transiciones: Diccionario de transiciones
        path_salida: Directorio de salida
    """
    print(f"\n4. Generando mapa de transición...")

    # Leer rasters
    with rasterio.open(path_clc) as src:
        CLC = src.read(1)
        meta = src.meta.copy()

    with rasterio.open(path_bau) as src:
        BaU = src.read(1)

    with rasterio.open(path_idq) as src:
        IDQ = src.read(1)

    with rasterio.open(path_portafolio) as src:
        SbN = src.read(1)

    print(f"  Dimensiones: {CLC.shape}")

    # Inicializar
    resultado = np.full_like(CLC, 1, dtype=np.int32)
    transicion_info = []
    degradaciones = {}
    next_trans_id = max([v[1] for v in tabla_transiciones.values()]) + 1

    con_nbs = 0
    con_deg = 0
    sin_cambio = 0

    rows, cols = CLC.shape

    print("  Procesando píxeles...")
    for i in range(rows):
        for j in range(cols):
            clc = CLC[i, j]
            bau = BaU[i, j]
            sbn = SbN[i, j]
            idq = IDQ[i, j]

            if clc == -9999 or bau == -9999 or sbn == -9999 or idq == -9999:
                continue

            clave = (int(clc), int(sbn))

            if clave in tabla_transiciones:
                # Transición NbS
                clc_end, trans_id, accion = tabla_transiciones[clave]
                resultado[i, j] = trans_id
                transicion_info.append((trans_id, clc_end, 1, idq, accion))
                con_nbs += 1
            else:
                if clc == bau:
                    resultado[i, j] = 1
                    sin_cambio += 1
                else:
                    clave_deg = (int(clc), int(bau))
                    if clave_deg not in degradaciones:
                        degradaciones[clave_deg] = next_trans_id
                        next_trans_id += 1
                    trans_id = degradaciones[clave_deg]
                    resultado[i, j] = trans_id
                    transicion_info.append((trans_id, bau, 1, idq, 7))
                    con_deg += 1

        if (i + 1) % 100 == 0:
            print(f"    Fila {i+1}/{rows} ({(i+1)/rows*100:.1f}%)")

    print(f"  NbS: {con_nbs}, Degradación: {con_deg}, Sin cambio: {sin_cambio}")

    # Guardar raster
    path_out = Path(path_salida) / 'transicion.tif'
    with rasterio.open(path_out, 'w', **meta) as dst:
        dst.write(resultado, 1)

    # Tabla de transiciones
    df_trans = pd.DataFrame(transicion_info, columns=["transicion", "CLC End", "deltatm", "idQ", "accion"])
    df_trans = df_trans.drop_duplicates()

    trans_positivas = pd.DataFrame([
        {"transicion": v[1], "CLC End": v[0], "deltatm": 1, "idQ": 1, "accion": v[2]}
        for v in tabla_transiciones.values()
    ])

    fila_trans_1 = pd.DataFrame([{"transicion": 1, "CLC End": -9999, "deltatm": 0, "idQ": 1, "accion": 1}])

    df_final = pd.concat([fila_trans_1, trans_positivas, df_trans], ignore_index=True)
    df_final = df_final.drop_duplicates(subset=["transicion"])
    df_final = df_final.sort_values(by="transicion").reset_index(drop=True)
    df_final.to_excel(Path(path_salida) / 'Tabla_transiciones_idQ.xlsx', index=False)

    print(f"  ✓ Transición guardada: {path_out.name}")


def generar_serie_temporal_BaU_simple(area_degradacion, path_salida, config):
    """
    Genera serie temporal para BaU (solo columna 5 con degradación)

    Args:
        area_degradacion: Área total de degradación en m²
        path_salida: Directorio de salida
        config: Objeto ConfiguracionEscenario
    """
    print(f"  Área de degradación: {area_degradacion:,.2f} m² ({area_degradacion/10000:,.2f} ha)")

    # Generar serie temporal
    fechas = pd.date_range(start=config.fecha_inicio, end=config.fecha_fin)
    df = pd.DataFrame(index=fechas)

    # Separar fecha en columnas Año, Mes, Día
    df['Año'] = df.index.year
    df['Mes'] = df.index.month
    df['Día'] = df.index.day

    # Inicializar todas las categorías (1, 2, 3, 4, 5) en 0
    categorias = [1, 2, 3, 4, 5]
    for cat in categorias:
        df[cat] = 0.0

    # Calcular distribución temporal
    start_year = pd.to_datetime(config.fecha_inicio).year
    anios_dist = list(range(start_year, start_year + config.anios_distribucion))
    dias_distribucion = sum(366 if is_leap_year(a) else 365 for a in anios_dist)

    # Solo la categoría 5 tiene área (BaU = todo se degrada)
    area_diaria = area_degradacion / dias_distribucion if dias_distribucion > 0 else 0

    print(f"  Días de distribución: {dias_distribucion} ({config.anios_distribucion} años)")
    print(f"  Área diaria: {area_diaria:,.2f} m²/día")
    print(f"  Área diaria: {area_diaria:.2e} m²/día")

    # Asignar valores solo a columna 5, resto en 0
    serie_cat5 = [area_diaria] * dias_distribucion + [0] * (len(df) - dias_distribucion)
    df[5] = serie_cat5[:len(df)]

    # Reordenar columnas
    df = df[['Año', 'Mes', 'Día', 1, 2, 3, 4, 5]]

    # Exportar CSV
    df.to_csv(Path(path_salida) / 'Serie_diaria_areas.csv', index=False)

    # Exportar formato SIGA (solo columnas numéricas)
    try:
        from common_functions.ficheros import Ficheros

        df_meta = pd.DataFrame({
            "codigo": categorias,
            "x": [-9999] * len(categorias),
            "y": [-9999] * len(categorias),
            "z": [-9999] * len(categorias)
        })

        dir_series = Path(path_salida) / 'series'
        dir_series.mkdir(exist_ok=True)

        # Crear DataFrame solo con columnas numéricas para SIGA
        df_siga = df[categorias].copy()

        ficheros = Ficheros(ruta_cuenca_SIGA=str(path_salida) + '/')
        ficheros.ExportarSeriesSIGA(
            df=df_siga,
            df_metadatos=df_meta,
            nombre_archivo_serie='/series_TasRestBaU.txt'
        )
        print(f"  ✓ Serie SIGA exportada")
    except ImportError:
        print(f"  ⚠ Módulo 'ficheros' no disponible - solo CSV exportado")
    except Exception as e:
        print(f"  ⚠ Error al exportar SIGA: {e}")

    print(f"  ✓ Serie temporal BaU generada")


def generar_serie_temporal_BaU(raster_portafolio, path_salida, config):
    """
    Genera serie temporal para BaU (solo columna 5 con degradación)

    Args:
        raster_portafolio: Raster Portafolio_Predios.tif
        path_salida: Directorio de salida
        config: Objeto ConfiguracionEscenario
    """
    with rasterio.open(raster_portafolio) as src:
        raster = src.read(1)
        resolution = src.res[0]
        nodata = src.nodata
        area_pixel = resolution * resolution

    # Calcular área de la categoría 5 (todas las áreas válidas en BaU)
    categorias_fijas = [1, 2, 3, 4, 5]
    valores_presentes = np.unique(raster[raster != nodata])

    areas_totales = {
        v: (raster == v).sum() * area_pixel
        for v in valores_presentes if v in categorias_fijas
    }

    # Asegurar que todas las categorías existan
    for c in categorias_fijas:
        areas_totales.setdefault(c, 0.0)

    print(f"  Áreas por categoría (m²):")
    for cat, area in areas_totales.items():
        if area > 0:
            print(f"    Cat {cat}: {area:,.2f} m² ({area/10000:,.2f} ha)")

    # Generar serie temporal
    fechas = pd.date_range(start=config.fecha_inicio, end=config.fecha_fin)
    df = pd.DataFrame(index=fechas)

    # Separar fecha en columnas Año, Mes, Día
    df['Año'] = df.index.year
    df['Mes'] = df.index.month
    df['Día'] = df.index.day

    # Inicializar todas las categorías (1, 2, 3, 4, 5)
    categorias = [1, 2, 3, 4, 5]
    for cat in categorias:
        df[cat] = 0.0

    # BaU: degradación distribuida en anios_distribucion (igual que código original)
    start_year = pd.to_datetime(config.fecha_inicio).year
    anios_dist = list(range(start_year, start_year + config.anios_distribucion))
    dias_distribucion = sum(366 if is_leap_year(a) else 365 for a in anios_dist)

    # Área diaria durante el periodo de distribución
    area_diaria = area_degradacion / dias_distribucion if dias_distribucion > 0 else 0

    print(f"  Días de distribución: {dias_distribucion} ({config.anios_distribucion} años)")
    print(f"  Área diaria: {area_diaria:,.2f} m²/día")

    # La columna 5 tiene valor durante dias_distribucion, luego 0
    serie_bau = [area_diaria] * dias_distribucion + [0] * (len(df) - dias_distribucion)
    df[5] = serie_bau[:len(df)]

    # Reordenar columnas
    df = df[['Año', 'Mes', 'Día', 1, 2, 3, 4, 5]]

    # Exportar CSV
    df.to_csv(Path(path_salida) / 'Serie_diaria_areas.csv', index=False)

    # Exportar formato SIGA (solo columnas numéricas)
    try:
        from common_functions.ficheros import Ficheros

        df_meta = pd.DataFrame({
            "codigo": categorias,
            "x": [-9999] * len(categorias),
            "y": [-9999] * len(categorias),
            "z": [-9999] * len(categorias)
        })

        dir_series = Path(path_salida) / 'series'
        dir_series.mkdir(exist_ok=True)

        # Crear DataFrame solo con columnas numéricas para SIGA
        df_siga = df[categorias].copy()

        ficheros = Ficheros(ruta_cuenca_SIGA=str(path_salida) + '/')
        ficheros.ExportarSeriesSIGA(
            df=df_siga,
            df_metadatos=df_meta,
            nombre_archivo_serie='/series_TasRestBaU.txt'
        )
        print(f"  ✓ Serie SIGA exportada")
    except ImportError:
        print(f"  ⚠ Módulo 'ficheros' no disponible - solo CSV exportado")
    except Exception as e:
        print(f"  ⚠ Error al exportar SIGA: {e}")

    print(f"  ✓ Serie temporal BaU generada")


def generar_serie_temporal(path_raster, start_date, end_date, distribucion_anios,
                           path_salida, nombre_escenario='Base'):
    """
    Genera serie temporal de áreas por categoría

    Args:
        path_raster: Raster con acciones (Portafolio_Predios.tif)
        start_date: Fecha inicio (YYYY-MM-DD)
        end_date: Fecha fin (YYYY-MM-DD)
        distribucion_anios: Años de implementación
        path_salida: Directorio de salida
        nombre_escenario: Nombre del escenario
    """
    print(f"\n5. Generando serie temporal...")

    categorias_fijas = [1, 2, 3, 4, 5]

    with rasterio.open(path_raster) as src:
        raster = src.read(1)
        resolution = src.res[0]
        nodata = src.nodata
        area_pixel = resolution * resolution

    # Calcular áreas totales
    valores_presentes = np.unique(raster[raster != nodata])
    areas_totales = {
        v: (raster == v).sum() * area_pixel
        for v in valores_presentes if v in categorias_fijas
    }

    for c in categorias_fijas:
        areas_totales.setdefault(c, 0.0)

    print(f"  Áreas por categoría (ha):")
    for cat, area in areas_totales.items():
        print(f"    Cat {cat}: {area/10000:,.2f} ha")

    # Generar fechas
    fechas = pd.date_range(start=start_date, end=end_date)
    df = pd.DataFrame(index=fechas)
    df.index.name = "Date"

    # Calcular días de distribución
    start_year = pd.to_datetime(start_date).year
    anios_dist = list(range(start_year, start_year + distribucion_anios))
    dias_dist = sum(366 if is_leap_year(a) else 365 for a in anios_dist)

    # Generar series
    for c in categorias_fijas:
        area_diaria = areas_totales[c] / dias_dist if dias_dist > 0 else 0
        serie = [area_diaria] * dias_dist + [0] * (len(df) - dias_dist)
        df[c] = serie[:len(df)]

    # Exportar CSV
    df.to_csv(Path(path_salida) / 'Serie_diaria_areas.csv', index=True)

    # Exportar SIGA
    try:
        from common_functions.ficheros import Ficheros

        n = len(categorias_fijas)
        df_meta = pd.DataFrame({
            "codigo": categorias_fijas,
            "x": [-9999] * n,
            "y": [-9999] * n,
            "z": [-9999] * n
        })

        ficheros = Ficheros(ruta_cuenca_SIGA=str(path_salida) + '/')
        ficheros.ExportarSeriesSIGA(
            df=df,
            df_metadatos=df_meta,
            nombre_archivo_serie=f'/series_TasRest{nombre_escenario}.txt'
        )
        print(f"  ✓ Serie SIGA exportada")
    except ImportError:
        print(f"  ⚠ Módulo 'ficheros' no disponible - solo CSV exportado")

    print(f"  ✓ Serie temporal generada")


# ==============================================================================
# MAIN - PUNTO DE ENTRADA
# ==============================================================================

if __name__ == "__main__":
    """
    INSTRUCCIONES DE USO:

    1. Editar la configuración abajo según tu escenario
    2. Ejecutar: python generador_escenarios_SIGA.py
    """

    # =========================================================================
    # CONFIGURACIÓN - EDITAR AQUÍ
    # =========================================================================

    # Seleccionar escenario: 'A', 'B', 'C'
    TIPO_ESCENARIO = 'B'

    # -------------------------------------------------------------------------
    # OPCIÓN 1: Usar configuración por defecto
    # -------------------------------------------------------------------------
    config = ConfiguracionEscenario(tipo_escenario=TIPO_ESCENARIO)

    # -------------------------------------------------------------------------
    # OPCIÓN 2: Personalizar parámetros al crear la configuración
    # -------------------------------------------------------------------------
    # Ejemplo con parámetros personalizados:
    mi_dict = {
        1: 1,   # Conservación
        2: 2,   # Reforestación activa
        3: 2,   # Reforestación pasiva (también mapea a acción 2)
        4: 3,   # Regeneración natural
        0: 4    # Sin acción
    }

    config = ConfiguracionEscenario(
        tipo_escenario=TIPO_ESCENARIO,
        dict_transformacion=mi_dict,      # Diccionario personalizado
        valor_fuera_poligono=5,           # Valor para fuera de polígonos
        anios_distribucion=5,            # Años de implementación NbS
        fecha_inicio='2025-01-01',        # Fecha inicio simulación
        fecha_fin='2075-12-31'            # Fecha fin simulación
    )

    # -------------------------------------------------------------------------
    # OPCIÓN 3: Modificar parámetros después de crear config
    # -------------------------------------------------------------------------
    # config.path_datos = "C:/Users/miguel.canon/Box/.../01_Data/"
    # config.path_salida = "C:/Users/miguel.canon/Box/.../02_OUT/"
    # config.montos_millones = [75, 100, 200]  # Cambiar presupuestos
    # config.anios_distribucion = 20           # Cambiar años

    # =========================================================================
    # EJECUTAR ESCENARIO
    # =========================================================================

    try:
        if TIPO_ESCENARIO == 'A':
            generar_escenario_A(config)

        elif TIPO_ESCENARIO == 'B':
            generar_escenario_B(config)

        elif TIPO_ESCENARIO == 'C':
            generar_escenario_C(config)

        elif TIPO_ESCENARIO == 'D':
            generar_escenario_D(config)

        else:
            print(f"ERROR: Escenario '{TIPO_ESCENARIO}' no reconocido")
            print("Opciones válidas: 'A', 'B', 'C', 'D'")

    except Exception as e:
        print(f"\n✗ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
